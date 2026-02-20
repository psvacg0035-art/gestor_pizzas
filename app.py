from flask import Flask, render_template, request, redirect, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)

# ============================
# 🔴 CONFIGURACIÓN SEGURA POSTGRESQL
# ============================

database_url = os.getenv("DATABASE_URL")

if not database_url:
    raise RuntimeError("DATABASE_URL no está configurada en Render")

if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ============================
# MODELO ORIGINAL (NO MODIFICADO)
# ============================

class Pedido(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente = db.Column(db.String(100))
    departamento = db.Column(db.String(100))
    telefono = db.Column(db.String(50))
    sabor = db.Column(db.String(100))
    cantidad = db.Column(db.Integer)
    precio = db.Column(db.Integer)
    total = db.Column(db.Integer)
    hora_entrega = db.Column(db.String(10))
    estado = db.Column(db.String(50), default="Pendiente")
    observaciones = db.Column(db.String(200))
    metodo_pago = db.Column(db.String(50))
    fecha = db.Column(db.Date, default=datetime.today)

# ============================
# 🆕 NUEVO MODELO PARA CONFIGURAR CAMPOS
# ============================

class ConfigCampo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    clave = db.Column(db.String(50), unique=True)
    nombre_mostrar = db.Column(db.String(100))
    activo = db.Column(db.Boolean, default=True)

# ============================
# Crear tablas y cargar configuración inicial
# ============================

with app.app_context():
    db.create_all()

    campos_base = [
        ("cliente", "Cliente"),
        ("departamento", "Departamento"),
        ("telefono", "Telefono"),
        ("sabor", "Sabor"),
        ("cantidad", "Cantidad"),
        ("precio", "Precio"),
    ]

    for clave, nombre in campos_base:
        existe = ConfigCampo.query.filter_by(clave=clave).first()
        if not existe:
            nuevo = ConfigCampo(clave=clave, nombre_mostrar=nombre)
            db.session.add(nuevo)

    db.session.commit()

# ============================
# RUTAS
# ============================

@app.route("/")
def index():
    hoy = datetime.today().date()

    pedidos_activos = Pedido.query.filter(
        Pedido.estado != "Entregado",
        Pedido.fecha == hoy
    ).all()

    pedidos_entregados = Pedido.query.filter(
        Pedido.estado == "Entregado",
        Pedido.fecha == hoy
    ).all()

    total_dia = sum(p.total for p in pedidos_activos)
    total_entregado = sum(p.total for p in pedidos_entregados)
    total_general = total_dia + total_entregado

    # 🆕 AGREGADO: enviar configuración a la vista
    config_campos = ConfigCampo.query.all()
    config_dict = {c.clave: c for c in config_campos}

    return render_template(
        "index.html",
        pedidos_activos=pedidos_activos,
        pedidos_entregados=pedidos_entregados,
        total_dia=total_dia,
        total_entregado=total_entregado,
        total_general=total_general,
        fecha_hoy=hoy,
        config=config_dict   # 🆕 agregado
    )

@app.route("/agregar", methods=["POST"])
def agregar():
    cliente = request.form["cliente"]
    departamento = request.form["departamento"]
    telefono = request.form["telefono"]
    sabor = request.form["sabor"]
    cantidad = int(request.form["cantidad"])
    precio = int(request.form["precio"])
    hora = request.form["hora"]
    observaciones = request.form.get("observaciones")

    total = cantidad * precio

    nuevo = Pedido(
        cliente=cliente,
        departamento=departamento,
        telefono=telefono,
        sabor=sabor,
        cantidad=cantidad,
        precio=precio,
        total=total,
        hora_entrega=hora,
        observaciones=observaciones
    )

    db.session.add(nuevo)
    db.session.commit()

    return redirect("/")

@app.route("/cambiar_estado/<int:id>/<estado>")
def cambiar_estado(id, estado):
    pedido = Pedido.query.get(id)
    if pedido:
        pedido.estado = estado
        db.session.commit()
    return redirect("/")

@app.route("/entregar", methods=["POST"])
def entregar():
    pedido_id = request.form.get("pedido_id")
    metodo_pago = request.form.get("metodo_pago")

    pedido = Pedido.query.get(pedido_id)

    if pedido:
        pedido.estado = "Entregado"
        pedido.metodo_pago = metodo_pago
        db.session.commit()

    return redirect("/")

@app.route("/eliminar/<int:id>")
def eliminar(id):
    pedido = Pedido.query.get(id)
    if pedido:
        db.session.delete(pedido)
        db.session.commit()
    return redirect("/")

# ============================
# HISTORIAL COMPLETO
# ============================

@app.route("/historial")
def historial():
    pedidos = Pedido.query.order_by(Pedido.fecha.desc(), Pedido.id.desc()).all()
    return render_template("historial.html", pedidos=pedidos)

# ============================
# EXPORTAR A EXCEL
# ============================

@app.route("/exportar_excel")
def exportar_excel():
    pedidos = Pedido.query.order_by(Pedido.fecha.asc(), Pedido.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Pedidos"

    encabezados = [
        "ID",
        "Fecha",
        "Cliente",
        "Departamento",
        "Telefono",
        "Sabor",
        "Cantidad",
        "Precio",
        "Total",
        "Estado",
        "Observaciones",
        "Metodo Pago"
    ]

    ws.append(encabezados)

    for p in pedidos:
        ws.append([
            p.id,
            p.fecha,
            p.cliente,
            p.departamento,
            p.telefono,
            p.sabor,
            p.cantidad,
            p.precio,
            p.total,
            p.estado,
            p.observaciones or "",
            p.metodo_pago or ""
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="historial_pedidos.xlsx",
        as_attachment=True
    )

# ============================
# 🆕 PANEL ADMINISTRAR CAMPOS
# ============================

@app.route("/admin_campos")
def admin_campos():
    campos = ConfigCampo.query.all()
    return render_template("admin_campos.html", campos=campos)

@app.route("/editar_campo/<int:id>", methods=["POST"])
def editar_campo(id):
    campo = ConfigCampo.query.get(id)
    if campo:
        campo.nombre_mostrar = request.form["nombre"]
        campo.activo = "activo" in request.form
        db.session.commit()
    return redirect("/admin_campos")

# 🔴 NO MODIFICADO
if __name__ == "__main__":
    app.run(debug=True)
